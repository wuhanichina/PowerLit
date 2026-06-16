from __future__ import annotations

import csv
from re import sub
from typing import Any

from powerlit.models import DocumentType, PaperRecord
from powerlit.settings import Settings

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


TITLE_HEADERS = (
    "journal_title",
    "source_title",
    "source title",
    "journal",
    "journal name",
    "期刊名称",
    "刊名",
    "刊名英文",
    "英文刊名",
    "刊名（英文）",
)
PREFERRED_QUARTILE_HEADERS = (
    "大类分区",
    "升级版大类分区",
    "基础版大类分区",
    "cas大类分区",
    "cas quartile",
    "cas_quartile",
    "quartile",
)
FALLBACK_QUARTILE_HEADERS = (
    "分区",
    "小类分区",
    "升级版小类分区",
    "基础版小类分区",
)


class CASWhitelistService:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.path = settings.cas_journal_list_path
        self.max_quartile = settings.cas_max_quartile
        self._allowed_titles: set[str] | None = None

    @property
    def enabled(self) -> bool:
        return self.path.exists()

    def filter_records(self, records: list[PaperRecord]) -> list[PaperRecord]:
        if not self.enabled:
            return records
        allowed_titles = self._load_allowed_titles()
        filtered: list[PaperRecord] = []
        for record in records:
            if record.document_type != DocumentType.JOURNAL:
                continue
            title = normalize_journal_title(record.source_title)
            if title in allowed_titles:
                filtered.append(record)
        return filtered

    def _load_allowed_titles(self) -> set[str]:
        if self._allowed_titles is not None:
            return self._allowed_titles

        rows = list(self._load_rows())
        allowed_titles: set[str] = set()
        for row in rows:
            title = pick_value(row, TITLE_HEADERS)
            quartile_text = pick_value(row, PREFERRED_QUARTILE_HEADERS)
            if quartile_text is None:
                quartile_text = pick_value(row, FALLBACK_QUARTILE_HEADERS)
            quartile = parse_quartile(quartile_text)
            normalized_title = normalize_journal_title(title)
            if normalized_title and quartile is not None and quartile <= self.max_quartile:
                allowed_titles.add(normalized_title)

        if not allowed_titles:
            raise ValueError(
                "CAS 白名单文件已找到，但没有解析出有效期刊。"
                "请检查表头是否包含期刊名和分区列。"
            )
        self._allowed_titles = allowed_titles
        return allowed_titles

    def _load_rows(self) -> list[dict[str, str]]:
        suffix = self.path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv_rows()
        if suffix in {".xlsx", ".xlsm"}:
            return self._load_xlsx_rows()
        raise ValueError(f"不支持的 CAS 白名单文件格式: {self.path.name}")

    def _load_csv_rows(self) -> list[dict[str, str]]:
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with self.path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    return [normalize_row(row) for row in reader if row]
            except UnicodeDecodeError:
                continue
        raise ValueError(f"无法读取 CAS 白名单文件编码: {self.path}")

    def _load_xlsx_rows(self) -> list[dict[str, str]]:
        if load_workbook is None:
            raise ValueError("读取 xlsx 白名单需要安装 openpyxl。")
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        payload: list[dict[str, str]] = []
        for row in rows[1:]:
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else None
                record[header] = "" if value is None else str(value).strip()
            if record:
                payload.append(normalize_row(record))
        return payload


def normalize_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        normalize_header(str(key)): ("" if value is None else str(value).strip())
        for key, value in row.items()
        if key is not None
    }


def pick_value(row: dict[str, str], aliases: tuple[str, ...]) -> str | None:
    for alias in aliases:
        value = row.get(normalize_header(alias))
        if value:
            return value
    return None


def parse_quartile(value: str | None) -> int | None:
    if not value:
        return None
    normalized = value.strip().lower()
    replacements = {
        "一区": "1区",
        "二区": "2区",
        "三区": "3区",
        "四区": "4区",
        "q1": "1区",
        "q2": "2区",
        "q3": "3区",
        "q4": "4区",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    for digit in ("1", "2", "3", "4"):
        if f"{digit}区" in normalized or normalized == digit:
            return int(digit)
    return None


def normalize_header(value: str) -> str:
    return sub(r"[^0-9a-z\u4e00-\u9fff]+", "", value.casefold())


def normalize_journal_title(value: str | None) -> str | None:
    if not value:
        return None
    normalized = sub(r"[^0-9a-z\u4e00-\u9fff]+", "", value.casefold())
    return normalized or None
